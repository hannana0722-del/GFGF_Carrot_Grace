from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import date, datetime
from typing import List, Optional
import io

from app.database import get_db
from app.models.models import (
    User, Attendance, HealthCheck, DailyReport,
    Risk, Assignment, AssignmentSubmission, Notification
)
from app.schemas.schemas import (
    KPIMetrics, AttendanceChartData, AlertItem, DashboardReport,
    HealthCheckStats, RiskDashboard, RiskResponse, UserSummary
)
from app.core.security import get_current_user, require_manager_or_admin

router = APIRouter(prefix="/dashboard", tags=["Dashboard"])


def _get_status_val(obj, attr):
    val = getattr(obj, attr, None)
    if val is None:
        return None
    return val.value if hasattr(val, 'value') else str(val)


@router.get("/kpi", response_model=KPIMetrics)
async def get_kpi_metrics(
    target_date: Optional[date] = None,
    cohort_number: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager_or_admin)
):
    """Get today's KPI metrics."""
    today = target_date or date.today()

    # Total participants
    participant_query = db.query(User).filter(
        User.role == "participant",
        User.is_active == True
    )
    if cohort_number:
        participant_query = participant_query.filter(User.cohort_number == cohort_number)
    elif current_user.role == "manager" and current_user.cohort_number:
        participant_query = participant_query.filter(User.cohort_number == current_user.cohort_number)

    total_participants = participant_query.count()
    participant_ids = [u.id for u in participant_query.all()]

    if total_participants == 0:
        return KPIMetrics(
            date=str(today),
            total_participants=0,
            attendance_rate=0.0,
            report_submission_rate=0.0,
            health_green_rate=0.0,
            active_risks=0,
            emergency_alerts=0,
            pending_assignments=0
        )

    # Attendance rate for today
    att_records = db.query(Attendance).filter(
        Attendance.session_date == today,
        Attendance.user_id.in_(participant_ids)
    ).all()
    present_count = sum(
        1 for r in att_records
        if _get_status_val(r, "status") in ("present", "late")
    )
    attendance_rate = round(present_count / total_participants * 100, 1) if total_participants > 0 else 0.0

    # Report submission rate
    reports_today = db.query(DailyReport).filter(
        DailyReport.report_date == today,
        DailyReport.user_id.in_(participant_ids)
    ).count()
    report_rate = round(reports_today / total_participants * 100, 1)

    # Health green rate
    health_checks_today = db.query(HealthCheck).filter(
        HealthCheck.check_date == today,
        HealthCheck.user_id.in_(participant_ids)
    ).all()
    green_count = sum(
        1 for h in health_checks_today
        if _get_status_val(h, "signal_color") == "green"
    )
    health_green_rate = round(green_count / len(health_checks_today) * 100, 1) if health_checks_today else 0.0

    # Active risks (red or yellow)
    active_risks = db.query(Risk).filter(
        Risk.user_id.in_(participant_ids),
        Risk.status.in_(["red", "yellow"])
    ).count()

    # Emergency alerts (health checks with emergency_requested = True, today)
    emergency_alerts = db.query(HealthCheck).filter(
        HealthCheck.check_date == today,
        HealthCheck.user_id.in_(participant_ids),
        HealthCheck.emergency_requested == True
    ).count()

    # Pending assignments (due today or overdue, with missing submissions)
    pending_assignments = db.query(Assignment).filter(
        Assignment.due_date <= datetime.combine(today, datetime.max.time()),
        Assignment.cohort_number == cohort_number if cohort_number else True
    ).count()

    return KPIMetrics(
        date=str(today),
        total_participants=total_participants,
        attendance_rate=attendance_rate,
        report_submission_rate=report_rate,
        health_green_rate=health_green_rate,
        active_risks=active_risks,
        emergency_alerts=emergency_alerts,
        pending_assignments=pending_assignments
    )


@router.get("/attendance-chart", response_model=AttendanceChartData)
async def get_attendance_chart(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    group_by: str = Query("country", enum=["country", "team"]),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager_or_admin)
):
    """Get attendance chart data grouped by country or team."""
    query = db.query(Attendance)

    if start_date:
        query = query.filter(Attendance.session_date >= start_date)
    if end_date:
        query = query.filter(Attendance.session_date <= end_date)

    records = query.all()
    users = {u.id: u for u in db.query(User).all()}

    # Group data
    groups = {}
    for r in records:
        user = users.get(r.user_id)
        if not user:
            continue
        key = user.country if group_by == "country" else user.team
        if not key:
            continue
        if key not in groups:
            groups[key] = {"present": 0, "late": 0, "absent": 0, "early_leave": 0}
        status = _get_status_val(r, "status") or "absent"
        groups[key][status] = groups[key].get(status, 0) + 1

    labels = sorted(groups.keys())
    colors = {
        "present": "#22c55e",
        "late": "#f59e0b",
        "absent": "#ef4444",
        "early_leave": "#8b5cf6"
    }

    datasets = [
        {
            "label": status_name.replace("_", " ").title(),
            "data": [groups.get(label, {}).get(status_name, 0) for label in labels],
            "backgroundColor": color
        }
        for status_name, color in colors.items()
    ]

    return AttendanceChartData(labels=labels, datasets=datasets)


@router.get("/alerts", response_model=List[AlertItem])
async def get_alerts(
    limit: int = 20,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager_or_admin)
):
    """Get recent alerts and notifications."""
    alerts = []
    today = date.today()

    # Red health checks
    red_checks = db.query(HealthCheck).filter(
        HealthCheck.signal_color == "red",
        HealthCheck.check_date == today
    ).limit(10).all()

    for check in red_checks:
        user = db.query(User).filter(User.id == check.user_id).first()
        alerts.append(AlertItem(
            id=check.id,
            type="health_red",
            message=f"Red health signal from {user.full_name or user.username}",
            priority="high",
            user=UserSummary.model_validate(user) if user else None,
            created_at=datetime.combine(check.check_date, datetime.min.time())
        ))

    # Emergency health requests
    emergency_checks = db.query(HealthCheck).filter(
        HealthCheck.emergency_requested == True,
        HealthCheck.check_date == today
    ).limit(5).all()

    for check in emergency_checks:
        user = db.query(User).filter(User.id == check.user_id).first()
        alerts.append(AlertItem(
            id=check.id,
            type="health_emergency",
            message=f"EMERGENCY request from {user.full_name or user.username}",
            priority="critical",
            user=UserSummary.model_validate(user) if user else None,
            created_at=datetime.combine(check.check_date, datetime.min.time())
        ))

    # High priority risks
    high_risks = db.query(Risk).filter(
        Risk.status == "red",
        Risk.priority == "high"
    ).order_by(Risk.updated_at.desc()).limit(10).all()

    for risk in high_risks:
        user = db.query(User).filter(User.id == risk.user_id).first()
        issue = _get_status_val(risk, "issue_type") or "unknown"
        alerts.append(AlertItem(
            id=risk.id,
            type="risk_high",
            message=f"High risk ({issue}) flagged for {user.full_name or user.username if user else 'Unknown'}",
            priority="high",
            user=UserSummary.model_validate(user) if user else None,
            created_at=risk.updated_at or risk.created_at
        ))

    # Sort by priority and date
    priority_order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    alerts.sort(key=lambda a: (priority_order.get(a.priority, 99), -a.created_at.timestamp()))

    return alerts[:limit]


@router.get("/report", response_model=DashboardReport)
async def get_dashboard_report(
    report_date: Optional[date] = None,
    cohort_number: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager_or_admin)
):
    """Generate comprehensive dashboard report data."""
    today = report_date or date.today()

    # Get KPI
    participant_query = db.query(User).filter(User.role == "participant", User.is_active == True)
    if cohort_number:
        participant_query = participant_query.filter(User.cohort_number == cohort_number)
    total_participants = participant_query.count()
    participant_ids = [u.id for u in participant_query.all()]

    att_records = db.query(Attendance).filter(
        Attendance.session_date == today,
        Attendance.user_id.in_(participant_ids)
    ).all()
    present_count = sum(
        1 for r in att_records
        if _get_status_val(r, "status") in ("present", "late")
    )
    att_rate = round(present_count / total_participants * 100, 1) if total_participants > 0 else 0.0

    reports_today = db.query(DailyReport).filter(
        DailyReport.report_date == today,
        DailyReport.user_id.in_(participant_ids)
    ).count()
    report_rate = round(reports_today / total_participants * 100, 1) if total_participants > 0 else 0.0

    health_today = db.query(HealthCheck).filter(
        HealthCheck.check_date == today,
        HealthCheck.user_id.in_(participant_ids)
    ).all()
    green_c = sum(1 for h in health_today if _get_status_val(h, "signal_color") == "green")
    yellow_c = sum(1 for h in health_today if _get_status_val(h, "signal_color") == "yellow")
    red_c = sum(1 for h in health_today if _get_status_val(h, "signal_color") == "red")
    total_hc = len(health_today)
    health_green_rate = round(green_c / total_hc * 100, 1) if total_hc > 0 else 0.0

    active_risks = db.query(Risk).filter(
        Risk.user_id.in_(participant_ids),
        Risk.status.in_(["red", "yellow"])
    ).count()

    emergency = db.query(HealthCheck).filter(
        HealthCheck.check_date == today,
        HealthCheck.user_id.in_(participant_ids),
        HealthCheck.emergency_requested == True
    ).count()

    kpi = KPIMetrics(
        date=str(today),
        total_participants=total_participants,
        attendance_rate=att_rate,
        report_submission_rate=report_rate,
        health_green_rate=health_green_rate,
        active_risks=active_risks,
        emergency_alerts=emergency,
        pending_assignments=0
    )

    # Health summary
    health_stats = HealthCheckStats(
        total=total_hc,
        green=green_c,
        yellow=yellow_c,
        red=red_c,
        emergency_count=emergency,
        green_pct=round(green_c / total_hc * 100, 1) if total_hc > 0 else 0.0,
        yellow_pct=round(yellow_c / total_hc * 100, 1) if total_hc > 0 else 0.0,
        red_pct=round(red_c / total_hc * 100, 1) if total_hc > 0 else 0.0
    )

    # Risk summary
    all_risks_q = db.query(Risk).filter(Risk.user_id.in_(participant_ids)).all()
    red_risks = [r for r in all_risks_q if _get_status_val(r, "status") == "red"]
    by_issue = {}
    for r in all_risks_q:
        it = _get_status_val(r, "issue_type") or "unknown"
        by_issue[it] = by_issue.get(it, 0) + 1

    recent_red = sorted(red_risks, key=lambda r: r.updated_at or r.created_at, reverse=True)[:5]
    risk_summary = RiskDashboard(
        total_risks=len(all_risks_q),
        red_count=len(red_risks),
        yellow_count=sum(1 for r in all_risks_q if _get_status_val(r, "status") == "yellow"),
        green_count=sum(1 for r in all_risks_q if _get_status_val(r, "status") == "green"),
        high_priority=sum(1 for r in all_risks_q if _get_status_val(r, "priority") == "high"),
        by_issue_type=by_issue,
        recent_red_risks=[RiskResponse.model_validate(r) for r in recent_red]
    )

    # Attendance summary by country
    users = {u.id: u for u in participant_query.all()}
    att_by_country = {}
    for r in att_records:
        user = users.get(r.user_id)
        if user and user.country:
            c = user.country
            if c not in att_by_country:
                att_by_country[c] = {"present": 0, "late": 0, "absent": 0, "total": 0}
            s = _get_status_val(r, "status") or "absent"
            att_by_country[c][s] = att_by_country[c].get(s, 0) + 1
            att_by_country[c]["total"] += 1

    return DashboardReport(
        generated_at=datetime.utcnow(),
        kpi=kpi,
        attendance_summary=att_by_country,
        health_summary=health_stats,
        risk_summary=risk_summary,
        top_risks=[RiskResponse.model_validate(r) for r in recent_red]
    )


@router.get("/export")
async def export_dashboard(
    report_date: Optional[date] = None,
    cohort_number: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager_or_admin)
):
    """Export dashboard data to Excel using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    today = report_date or date.today()

    wb = openpyxl.Workbook()

    # ── Style helpers ──────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header_row(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # ── Sheet 1: KPI Summary ──────────────────────────────────────────────────
    ws_kpi = wb.active
    ws_kpi.title = "KPI Summary"

    participant_query = db.query(User).filter(User.role == "participant", User.is_active == True)
    if cohort_number:
        participant_query = participant_query.filter(User.cohort_number == cohort_number)
    total_participants = participant_query.count()
    participant_ids = [u.id for u in participant_query.all()]

    att_records = db.query(Attendance).filter(
        Attendance.session_date == today,
        Attendance.user_id.in_(participant_ids)
    ).all()
    present_count = sum(
        1 for r in att_records
        if (_get_status_val(r, "status") or "") in ("present", "late")
    )
    reports_today = db.query(DailyReport).filter(
        DailyReport.report_date == today,
        DailyReport.user_id.in_(participant_ids)
    ).count()
    health_today = db.query(HealthCheck).filter(
        HealthCheck.check_date == today,
        HealthCheck.user_id.in_(participant_ids)
    ).all()
    green_count = sum(1 for h in health_today if _get_status_val(h, "signal_color") == "green")

    ws_kpi.append(["GSDF Dashboard Report", str(today)])
    ws_kpi.append([])
    ws_kpi.append(["Metric", "Value"])
    style_header_row(ws_kpi, 3, 2)

    kpi_rows = [
        ("Report Date", str(today)),
        ("Total Participants", total_participants),
        ("Attendance Rate", f"{round(present_count/total_participants*100,1) if total_participants else 0}%"),
        ("Daily Report Submission Rate", f"{round(reports_today/total_participants*100,1) if total_participants else 0}%"),
        ("Health Green Rate", f"{round(green_count/len(health_today)*100,1) if health_today else 0}%"),
        ("Emergency Alerts", sum(1 for h in health_today if h.emergency_requested)),
    ]
    for row in kpi_rows:
        ws_kpi.append(list(row))

    auto_width(ws_kpi)

    # ── Sheet 2: Attendance ───────────────────────────────────────────────────
    ws_att = wb.create_sheet("Attendance")
    ws_att.append(["User ID", "Name", "Country", "Team", "Date", "Session", "Status", "Check-In", "Check-Out", "Method"])
    style_header_row(ws_att, 1, 10)

    all_att = db.query(Attendance).filter(
        Attendance.session_date == today,
        Attendance.user_id.in_(participant_ids)
    ).all()
    users_map = {u.id: u for u in participant_query.all()}

    for r in all_att:
        user = users_map.get(r.user_id)
        ws_att.append([
            r.user_id,
            user.full_name if user else "",
            user.country if user else "",
            user.team if user else "",
            str(r.session_date),
            r.session_type or "",
            _get_status_val(r, "status") or "",
            str(r.check_in_time) if r.check_in_time else "",
            str(r.check_out_time) if r.check_out_time else "",
            _get_status_val(r, "method") or ""
        ])
    auto_width(ws_att)

    # ── Sheet 3: Health Checks ────────────────────────────────────────────────
    ws_hc = wb.create_sheet("Health Checks")
    ws_hc.append(["User ID", "Name", "Country", "Date", "Survey Type", "Signal", "Emergency", "Notes"])
    style_header_row(ws_hc, 1, 8)

    for h in health_today:
        user = users_map.get(h.user_id)
        ws_hc.append([
            h.user_id,
            user.full_name if user else "",
            user.country if user else "",
            str(h.check_date),
            _get_status_val(h, "survey_type") or "",
            _get_status_val(h, "signal_color") or "",
            "Yes" if h.emergency_requested else "No",
            h.notes or ""
        ])
    auto_width(ws_hc)

    # ── Sheet 4: Risks ────────────────────────────────────────────────────────
    ws_risk = wb.create_sheet("Risks")
    ws_risk.append(["Risk ID", "User ID", "Name", "Issue Type", "Priority", "Status", "Description", "Last Updated"])
    style_header_row(ws_risk, 1, 8)

    risks = db.query(Risk).filter(Risk.user_id.in_(participant_ids)).all()
    all_users_map = {u.id: u for u in db.query(User).all()}

    for r in risks:
        user = all_users_map.get(r.user_id)
        ws_risk.append([
            r.id,
            r.user_id,
            user.full_name if user else "",
            _get_status_val(r, "issue_type") or "",
            _get_status_val(r, "priority") or "",
            _get_status_val(r, "status") or "",
            r.description or "",
            str(r.updated_at or r.created_at)
        ])
    auto_width(ws_risk)

    # ── Sheet 5: Daily Reports ────────────────────────────────────────────────
    ws_dr = wb.create_sheet("Daily Reports")
    ws_dr.append(["User ID", "Name", "Country", "Report Date", "Emotion", "Satisfaction Score", "Activities"])
    style_header_row(ws_dr, 1, 7)

    dr_today = db.query(DailyReport).filter(
        DailyReport.report_date == today,
        DailyReport.user_id.in_(participant_ids)
    ).all()

    for dr in dr_today:
        user = all_users_map.get(dr.user_id)
        ws_dr.append([
            dr.user_id,
            user.full_name if user else "",
            user.country if user else "",
            str(dr.report_date),
            dr.emotion_tag or "",
            dr.satisfaction_score or "",
            (dr.activities or "")[:200]
        ])
    auto_width(ws_dr)

    # ── Save to bytes ─────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"gsdf_report_{today}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
